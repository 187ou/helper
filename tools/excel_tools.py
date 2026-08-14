"""Excel 工具（基于 openpyxl，真实可用）。"""
import logging
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


def read_excel(path: str) -> list[list[Any]]:
    """读取 Excel 返回二维数据（含表头）。"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        data = [list(row) for row in ws.iter_rows(values_only=True)]
        wb.close()
        logger.info("读取 Excel: %s (%d 行)", path, len(data))
        return data
    except Exception as e:
        logger.error("读取 Excel 失败 %s: %s", path, e)
        return []


def write_excel(path: str, data: list[list[Any]]) -> bool:
    """写入 Excel。"""
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        for row in data:
            ws.append(row)
        Path(path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        logger.info("写入 Excel: %s (%d 行)", path, len(data))
        return True
    except Exception as e:
        logger.error("写入 Excel 失败 %s: %s", path, e)
        return False


def merge_excel(files: list[str], out_path: str) -> dict[str, Any]:
    """合并多个 Excel（纵向拼接）。"""
    try:
        from openpyxl import Workbook, load_workbook
        out_wb = Workbook()
        out_ws = out_wb.active
        total = 0
        header = None
        for f in files:
            wb = load_workbook(f, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                wb.close()
                continue
            if header is None:
                header = rows[0]
                out_ws.append(header)
                start = 1
            else:
                start = 1 if rows[0] == header else 0
            for row in rows[start:]:
                out_ws.append(row)
                total += 1
            wb.close()
        Path(out_path).parent.mkdir(parents=True, exist_ok=True)
        out_wb.save(out_path)
        logger.info("合并 %d 个 Excel → %s (%d 行)", len(files), out_path, total)
        return {"files": files, "output": out_path, "rows": total}
    except Exception as e:
        logger.error("合并 Excel 失败: %s", e)
        return {"files": files, "output": out_path, "rows": 0, "error": str(e)}


def clean_blank_rows(path: str) -> int:
    """清理空白行，返回清除数量。"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb.active
        removed = 0
        for row in reversed(list(ws.iter_rows())):
            if all(cell.value is None for cell in row):
                ws.delete_rows(row[0].row)
                removed += 1
        wb.save(path)
        logger.info("清理空白行: %s (移除 %d 行)", path, removed)
        return removed
    except Exception as e:
        logger.error("清理空白行失败 %s: %s", path, e)
        return 0


def sheet_names(path: str) -> list[str]:
    """获取所有工作表名。"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception:
        return []
